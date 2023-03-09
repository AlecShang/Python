#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   mt.py
@Time    :   2020/01/29 11:46:11
@Author  :   Alec Shang
@Version :   1.0
@Contact :   shangjingweiwork@163.com
@License :   (C)Copyright 2019-2020
@Desc    :   None
'''
#######################


import openpyxl
import re


def createExcel(excelName, sheetName):
    # excel表的表名称
    # excelName = 't1' + '.xlsx'
    # 工作簿的名称
    # sheetName = '1月份'
    test = []
    time = ''
    pattern = re.compile(
        '(\d{4}[\.\-/年]{1}\d{1,2}[\.\-/月]{1}\d{1,2}[ ]{0,1}\d{2}[:]\d{2}[:]\d{2}).*([\u4E00-\u9FA5]+[A-Z0-9]{6}).*')
    book = openpyxl.load_workbook('./MT/' + excelName)
    sheet = book[sheetName]
    cells_rows = sheet.max_row
    cells_col = sheet.max_column
    cells = sheet.values  # 读取整个sheet
    for col in range(cells_col):
        # if cell!='':
        for row in range(cells_rows):
            b = re.findall(pattern, str(
                sheet.cell(row + 1, col + 1).value))
            if b != [] and time != b[0][0]:
                if '放行' in str(sheet.cell(row + 1, col + 1).value):
                    test.append(b[0][0])
                    test.append(b[0][1])
                    test.append('出口开启')
                    time = b[0][0]
                # if '识别车牌' in str(sheet.cell(row+1,col+1).value):
                if sheet.cell(row + 1, col + 1).value[0] == '0':
                    test.append(b[0][0])
                    test.append(b[0][1])
                    test.append('入口开启')
                    time = b[0][0]

    # 创建excel文档
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet['A1'] = '时间'
    sheet['B1'] = '车牌号'
    sheet['C1'] = '进离操作'
    x = 2
    for i in range(0, len(test), 3):
        sheet['A' + str(x)] = test[i]
        sheet['B' + str(x)] = test[i + 1]
        sheet['C' + str(x)] = test[i + 2]
        x += 1

    wb.save('./MT/' + sheetName + '.xlsx')
    return('./MT/' + sheetName + '.xlsx')


def loadExcel(excelName):
    book = openpyxl.load_workbook('./MT/' + excelName)
    return book.sheetnames
