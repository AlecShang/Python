#!/usr/local/bin/python3
# -*- encoding: utf-8 -*-
'''
@File    :   excel.py
@Time    :   2019/03/16 10:43:05
@Author  :   Alec Shang
@Version :   1.0
@Contact :   shangjingweiwork@163.com
@License :   (C)Copyright 2019-2020
@Desc    :   None
'''
#######################

# 读取excel的内容,并写入到新的excel中去

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 读取xlsx文件,即工作簿
wb = load_workbook(filename=r'312.xlsx')
# 获取sheet表
table_name = wb['Sheet1']
# l = []
# # 首先获取sheet表内的最大行和最大列,根据最大值读取数据
# try:
#     for i in range(table_name.max_row):
#         for cell in list(table_name.rows)[i]:
#             # print(cell.value)
#             l.append(cell.value)
# finally:
#     print(l)
#     print('Done...')


l = []

# 读取指定矩形区域内的内容
for cell in table_name["A7":"T16"]:
    for cells in cell:
        # print(cells.value)
        l.append(cells.value)
print(l)

# # 实例化workbook,即激活工作簿
# wb1 = Workbook()
# # 定义一个xlsx文件
# excel_name_new = 'total.xlsx'
# # 激活workbook某个工作表
# wb1_active = wb1.active
# # 定义激活工作表的名称title
# wb1_active.title = 'shangjingwei'
# # 写入数据
# for i in range(10):
#     #0-19  20-39 40-59
#     wb1_active.append(l[i*20:(1+i)*20])
# # 最后保存工作簿
# wb1.save(filename = excel_name_new)


# print(table_list)
# print(table_name['D7'].value)

# 将读取到的内容写入到新的excel当中
# from openpyxl import Workbook
# from openpyxl.utils import get_column_letter

# excel_total_data = Workbook()
# excel_name = 'total.xlsx'
# excel_active = excel_total_data.active
# excel_active.title = 'shang'
# excel_active.append([1,2,3])
# excel_total_data.save(filename = excel_name)
# print('Done......')
