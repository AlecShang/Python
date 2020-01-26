#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   kunMing.py
@Time    :   2019/12/10 11:48:28
@Author  :   Alec Shang
@Version :   1.0
@Contact :   shangjingweiwork@163.com
@License :   (C)Copyright 2019-2020
@Desc    :   None
'''
#######################


# 读取excel的内容,并写入到新的excel中去

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# 读取xlsx文件,即工作簿
wb = load_workbook(
    filename=r'/Users/shangjingwei/Documents/昆明市住房调研/调研统计/工作簿2-2.xlsx')


# 获取sheet表
table_name = wb['工作表1']

excel_total_data = Workbook()
excel_name = '第16题.xlsx'
excel_active = excel_total_data.active
excel_active.title = 'shang'

# 读取指定矩形区域内的内容
l = []
a = 0
b = 0
c = 0
d = 0
e = 0
f = 0
g = 0
for cell in table_name["m2":"m520"]:
    for cells in cell:
        if cells.value == None:
            l.append('')
            continue
        if '1' in cells.value:
            a += 1
        if '2' in cells.value:
            b += 1
        if '3' in cells.value:
            c += 1
        if '4' in cells.value:
            d += 1
        if '5' in cells.value:
            e += 1
        if '6' in cells.value:
            f += 1
        if '7' in cells.value:
            g += 1
        else:
            l.append('')
            continue
        # print(cells.value)
    # excel_active.append(l)
    # excel中单元格为B2开始，即第2列，第2行
    # for i in range(len(l)):
    #     # 判断是否应该换列
    #     excel_active.cell(i+2, 2).value=l[i]
    #     excel_total_data.save(filename = excel_name)

# print(l)
print(a, b, c, d, e, f, g)
print(round(a / 520, 4), round(b / 520, 4), round(c / 520, 4),
      round(d / 520, 4), round(e / 520, 4), round(f / 520, 4), round(g / 520, 4))
